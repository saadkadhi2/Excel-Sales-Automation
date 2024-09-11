import pandas as pd
import openpyxl
from openpyxl.styles import PatternFill
import matplotlib.pyplot as plt

df = pd.read_excel('SalesData.xlsx')
sales_by_region = df.groupby('Region')['Sales Amount'].sum().reset_index()
sales_by_salesperson = df.groupby('Salesperson')['Sales Amount'].sum().reset_index()


def highlight_sales(file_path, amount):
    wb = openpyxl.load_workbook(file_path)
    sheet = wb.active

    # for each row in col 5
    for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=5, max_col=5):
        for cell in row:
            if cell.value > amount:
                sheet[cell.coordinate].fill = PatternFill(fgColor="FFFF00", fill_type="solid")

    wb.save('highlighted_data.xlsx')


def plot_sales_by_region():
    plt.figure(figsize=(8, 8))
    plt.bar(sales_by_region['Region'], sales_by_region['Sales Amount'], color='red')
    plt.title('Sales by Region')
    plt.xlabel('Region')
    plt.ylabel('Sales Amount')
    plt.savefig('sales_by_region.png')
    plt.show()


def plot_sales_by_salesperson():
    plt.figure(figsize=(8, 8))
    plt.bar(sales_by_salesperson['Salesperson'], sales_by_salesperson['Sales Amount'], color='blue')
    plt.title('Sales by Salesperson')
    plt.xlabel('Salesperson')
    plt.ylabel('Sales Amount')
    plt.savefig('sales_by_salesperson.png')
    plt.show()


with pd.ExcelWriter('sales_summary.xlsx', engine='openpyxl') as writer:
    sales_by_region.to_excel(writer,
                             sheet_name='Sales by Region', index=False)
    sales_by_salesperson.to_excel(writer,
                                  sheet_name='Sales by Salesperson', index=False)

highlight_sales('SalesData.xlsx', 3000)
plot_sales_by_region()
plot_sales_by_salesperson()
